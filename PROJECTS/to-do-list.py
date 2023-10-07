#Reference : https://www.javatpoint.com/python-openpyxl

import time
from openpyxl import Workbook,load_workbook

# Define Colors for Opening message
red = "\033[38;2;255;0;0m"
green = "\033[38;2;0;255;0m"
reset_color = "\033[0m"

# Printing Opening Message
print(red + '''
 _____           ____          _     _     _   _ 
|_   _|__       |  _ \  ___   | |   (_)___| |_| |
  | |/ _ \ _____| | | |/ _ \  | |   | / __| __| |
  | | (_) |_____| |_| | (_) | | |___| \__ \ |_|_|
  |_|\___/      |____/ \___/  |_____|_|___/\__(_)

      ''')
print(green + "   By : Ash-codes18" + reset_color)
print("\n")

file_name = 'To-Do_List.xlsx'

try:
    # Try to open the existing Excel file
    workbook = load_workbook(file_name)
    sheet = workbook.active
except FileNotFoundError:
    # If the file doesn't exist, create it and add a new sheet
    workbook = Workbook()
    sheet = workbook.active
    # Initialize sheet headers
    sheet['A1'] = "Task"
    sheet['B1'] = "State"


#To display the list's data
def display():
    if sheet.max_row > 1:
        print("Your To-Do List : ")
        print("S.no            Tasks                                             States")
        for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
            task, state = row
            print(f"{i}.            {task}                                             {state}")
        print("\n")
    else:
        print("Your To-Do List is empty!\n")


#to change the status of a task
def change_status():
    display()
    while True:
        choice = int(input("Enter the task number you want to Change Status of : "))
        print()
        if 1 <= choice < sheet.max_row:
            print("States :\n1. Not Done\n2. In Progress\n3. Done\n")
            while True:
                state_val = int(input("Enter the state of the task : "))
                print()
                if 1 <= state_val <= 3:
                    sheet.cell(row=choice + 1, column=2, value=["Not Done", "In Progress", "Done"][state_val - 1])
                    workbook.save(file_name)
                    print("Status updated successfully!\n")
                    break
                else:
                    print("Invalid Choice, Please input the correct choice...\n")
            break
        else:
            print("Invalid Choice, Please input the correct choice...\n")

#to add new tasks to the list
def new_tasks():
    n = int(input("Enter the number of items you want to add : "))
    print()
    for i in range(n):
        print(f"Enter task {i + 1}: ", end='')
        task_name = input()
        print("States :\n1. Not Done\n2. In Progress\n3. Done\n")
        while True:
            state_val = int(input("Enter the state of the task : "))
            if 1 <= state_val <= 3:
                sheet.append([task_name, ["Not Done", "In Progress", "Done"][state_val - 1]])
                workbook.save(file_name)
                print("Task added successfully!\n")
                break
            else:
                print("Invalid Choice, Please input the correct choice...\n")

#to delete a task from the list
def del_item():
    display()
    choice = int(input("\nEnter the task number you want to delete : "))
    print()
    if 1 <= choice < sheet.max_row:
        sheet.delete_rows(choice + 1)
        workbook.save(file_name)
        print("Task deleted successfully!\n")
    else:
        print("Invalid Choice, Please input the correct choice...\n")
        

#To clear the list     
def empty_list():
    confirmation = input("Are you sure you want to empty the to-do list? (yes/no): ")
    if confirmation.lower() == "yes":
        sheet.delete_rows(2, sheet.max_row)  # Clear all rows except headers
        workbook.save(file_name)
        print("To-do list is now empty.")
    else:
        print("Operation canceled.")


while True:
    print("Options :\n1. Display List\n2. Add new item\n3. Change Status\n4. Delete item\n5. Delete List\n6. Quit\n")

    choice = int(input("Enter your choice : "))
    print()
    
    if choice == 1:
        display()

    elif choice == 2:
        new_tasks()

    elif choice == 3:
        change_status()

    elif choice == 4:
        del_item()

    elif choice == 5:
        empty_list()

    elif choice == 6:
        print("Exiting the app...\n")
        workbook.close()
        time.sleep(1)
        break

    else:
        print("\nInvalid Choice, Please input the correct choice...\n")
